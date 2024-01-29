import datetime

import decouple
import pandas as pd
import requests
from openpyxl import load_workbook
from pandas import DataFrame

import browser
from exceptions import FileNotExpectedError

MANDATORY_FIELDS = [
    "Date_Saved",
    "Employment_Options",
    "Work_Type",
    "Company",
    "Location",
    "Work_Type",
]
GOOGLE_PLACES_API_HEADERS = {
    "Content-Type": "application/json",
    "X-Goog-Api-Key": decouple.config("PLACES_API_KEY"),
    "X-Goog-FieldMask": "places.formattedAddress,places.addressComponents,places.location",
}


def check_nan(row, field):
    """
    Check if a dataframe value is NaN and return a tuple with the result and the field name.
    """
    return pd.isna(getattr(row, field)), field


def read_job_tracking_system(file_path):
    """
    Read the Job Tracking System.
    """
    try:
        job_tracking_system: DataFrame = pd.read_excel(
            file_path,
            sheet_name="Job Search Tracker",
            dtype={
                "Hbtn_Job_ID": "Int64",
                "Job_Position": "str",
                "Company": "str",
                "Work_Type": "str",
                "Employment_Options": "str",
                "Min_Salary": "Float64",
                "Max_Salary": "Float64",
                "Currency": "str",
                "Frequency": "str",
                "Location": "str",
                "Status": "str",
                "Excitement": "Int64",
                "Date_Saved": "datetime64[ns]",
                "Last_Updated": "datetime64[ns]",
                "Last_Uploaded": "datetime64[ns]",
                "Delete": "str",
            },
        )
        job_tracking_system_notes: DataFrame = pd.read_excel(
            file_path,
            sheet_name="Job Search Notes",
            dtype={"Hbtn_Job_ID": "Int64", "Note": "str", "Uploaded": "str"},
        )
    except pd.errors.ParserError:
        raise FileNotExpectedError("Wrong format for files")

    print("Job Tracking System successfully loaded!")
    print("Job Tracking System Notes successfully loaded!")
    return job_tracking_system, job_tracking_system_notes


def save_job_status_on_job_tracking_system(
    job_tracking_system, job_tracking_system_notes, file_path
):
    """
    Update the Job Tracking System
    """
    book = load_workbook(file_path)

    jts_ws = book["Job Search Tracker"]
    start_cell = jts_ws["A2"]
    for row in job_tracking_system.itertuples():
        for col_idx, value in enumerate(row[1:]):
            if pd.isnull(value):
                value = ""
            jts_ws.cell(
                row=start_cell.row,
                column=start_cell.column + col_idx,
                value=value,
            )
        start_cell = jts_ws.cell(row=start_cell.row + 1, column=start_cell.column)

    jts_notes_ws = book["Job Search Notes"]
    start_cell = jts_notes_ws["A2"]
    for row in job_tracking_system_notes.itertuples():
        for col_idx, value in enumerate(row[1:]):
            if pd.isnull(value):
                value = ""
            jts_notes_ws.cell(
                row=start_cell.row,
                column=start_cell.column + col_idx,
                value=value,
            )
        start_cell = jts_notes_ws.cell(row=start_cell.row + 1, column=start_cell.column)

    book.save(file_path)
    book.close()

    print("Job Tracking System successfully updated!")
    print("Job Tracking System Notes successfully updated!")


def check_job_data_integrity(row):
    """
    Check if there are mandatory fields from the Intranet that are missing in the Job Tracking System
    """
    results = [check_nan(row, field) for field in MANDATORY_FIELDS]
    if any(result[0] for result in results):
        print(
            f"No data in row {row.Index}. Missing data in columns:",
            ", ".join([f"{result[1]}" for result in results if result[0]]),
            "- Can't save form entry with incomplete data",
        )
        print(f"  Row data for debugging in your Job Tracking System: {row._asdict()}")
        return False
    return True


def fill_form_data(row):
    """
    Create an object with the form data needed to fill.
    """
    mandatory_fields_present = check_job_data_integrity(row)
    if not mandatory_fields_present:
        return None
    google_places_api_data = {"textQuery": row.Location}
    response = requests.post(
        url="https://places.googleapis.com/v1/places:searchText",
        json=google_places_api_data,
        headers=GOOGLE_PLACES_API_HEADERS,
    )
    parsed_response = response.json()
    place = parsed_response["places"][0]

    components = {}
    for component in place["addressComponents"]:
        for type in component["types"]:
            components[type] = component["shortText"]
    location_lat = place["location"]["latitude"]
    location_lng = place["location"]["longitude"]
    location_city = components.get("locality")
    location_state = components.get("administrative_area_level_1")
    location_country = components.get("country")

    end_date_year = ""
    end_date_month = ""
    end_date_day = ""
    if not pd.isna(row.Last_Updated) and row.Status in [
        "Declined",
        "Rejected",
        "Resigned",
        "Laid off",
    ]:
        end_date_year = row.Last_Updated.year
        end_date_month = row.Last_Updated.month
        end_date_day = row.Last_Updated.day

    salary = row.Min_Salary if not pd.isna(row.Min_Salary) else ""
    salary_currency = row.Currency if not pd.isna(row.Currency) else ""
    salary_frequency = row.Frequency if not pd.isna(row.Frequency) else ""

    form_data = {
        "#user_working_status_start_date_1i": str(row.Date_Saved.year),
        "#user_working_status_start_date_2i": str(row.Date_Saved.month),
        "#user_working_status_start_date_3i": str(row.Date_Saved.day),
        "#user_working_status_end_date_1i": str(end_date_year),
        "#user_working_status_end_date_2i": str(end_date_month),
        "#user_working_status_end_date_3i": str(end_date_day),
        "#user_working_status_employment": row.Employment_Options,
        "#user_working_status_work_type": row.Work_Type,
        "#company_name": row.Company,
        "#user_working_status_location_city": location_city,
        "#user_working_status_location_state": location_state,
        "#user_working_status_location_country": location_country,
        "#user_working_status_location_lat": location_lat,
        "#user_working_status_location_lng": location_lng,
        "#title": row.Job_Position,
        "#user_working_status_salary": str(salary),
        "#user_working_status_salary_currency": salary_currency,
        "#user_working_status_salary_frequency": salary_frequency,
    }
    return form_data


def select_job_notes(job_tracking_system_notes, job_id):
    """
    Select the job notes from the job to upload them
    """
    job_notes = job_tracking_system_notes.loc[
        (job_tracking_system_notes["Hbtn_Job_ID"] == job_id)
        & (job_tracking_system_notes["Uploaded"].isnull())
    ]
    return job_notes


def update_job_notes_status(job_tracking_system_notes, job_id):
    """
    Update the uploaded status from the job notes of the job tracking system.
    """
    job_tracking_system_notes.loc[
        (job_tracking_system_notes["Hbtn_Job_ID"] == job_id)
        & (job_tracking_system_notes["Uploaded"].isnull()),
        "Uploaded",
    ] = "Yes"


def process_jobs(job_tracking_system, job_tracking_system_notes, intranet):
    print("Processing jobs...")
    jobs_total = job_tracking_system.shape[0]
    jobs_uploaded = jobs_updated = jobs_deleted = 0
    for job in job_tracking_system.itertuples():
        # Delete job
        if job.Delete == "Yes":
            browser.delete_job_status(intranet, job)
            jobs_deleted += 1
            continue

        # Edit job
        if (
            not check_nan(job, "Last_Uploaded")[0]
            and job.Last_Updated > job.Last_Uploaded
        ):
            job_form_data = fill_form_data(job)
            if job_form_data is None:
                continue
            job_notes = select_job_notes(job_tracking_system_notes, job.Hbtn_Job_ID)
            browser.edit_job_status(intranet, job, job_form_data, job_notes)
            job_tracking_system.at[job.Index, "Last_Uploaded"] = pd.Timestamp.now()
            if job_notes.shape[0] != 0:
                update_job_notes_status(job_tracking_system_notes, job.Hbtn_Job_ID)
            jobs_updated += 1
            continue

        # Create new job
        if not check_nan(job, "Last_Uploaded")[0]:
            print(
                f"Job {job.Job_Position} from company {job.Company} already uploaded!"
            )
            continue
        job_form_data = fill_form_data(job)
        if job_form_data is None:
            continue
        job_notes = select_job_notes(job_tracking_system_notes, job.Hbtn_Job_ID)
        job_id = browser.create_job_status(intranet, job, job_form_data, job_notes)
        job_tracking_system.at[job.Index, "Last_Uploaded"] = pd.Timestamp.now()
        job_tracking_system.at[job.Index, "Hbtn_Job_ID"] = job_id
        jobs_uploaded += 1

    print(f"{jobs_uploaded} of {jobs_total} jobs successfully uploaded!")
    print(f"{jobs_updated} of {jobs_total} jobs successfully updated!")
    print(f"{jobs_deleted} of {jobs_total} jobs successfully deleted!")
